from __future__ import absolute_import
from __future__ import division, print_function, unicode_literals

from sumy.parsers.html import HtmlParser
from sumy.parsers.plaintext import PlaintextParser
from sumy.nlp.tokenizers import Tokenizer
from sumy.summarizers.lsa import LsaSummarizer as Summarizer
from sumy.nlp.stemmers import Stemmer
from sumy.utils import get_stop_words
from openpyxl import load_workbook
from bs4 import BeautifulSoup
import urllib
import urllib2

import requests
import re
import json as m_json
import sys

# pass in self as a parameter to each function is like doing this.get_videos() in c++
class getContent():

# the __init__ function is like a default contructor
    def __init__(self):
        print ("created a getContent object")
 
    def get_smry(self, input):
        smry_list = {}
        LANGUAGE = "english"
        SENTENCES_COUNT = 10
        parser = PlaintextParser.from_file(input, Tokenizer(LANGUAGE))
        stemmer = Stemmer(LANGUAGE)
    
        summarizer = Summarizer(stemmer)
        summarizer.stop_words = get_stop_words(LANGUAGE)

        i = 0
        for sentence in summarizer(parser.document, SENTENCES_COUNT):
            print(sentence)
            smry_list[str(i)] = str(sentence)
            i = i + 1
        return smry_list


    def get_videos(self, input):
        print ("get_videos() entered")
        list_videos = {}
        query = urllib.quote(input)
        url = "https://www.youtube.com/results?search_query=" + query
        response = urllib2.urlopen(url)
        html = response.read()
        soup = BeautifulSoup(html)
        i = 0
        print ("about to iterate through videos found")
        for vid in soup.findAll(attrs={'class':'yt-uix-tile-link'}):
            list_videos[str(i)] = 'https://www.youtube.com' + vid['href']
            i = i + 1
            print ("video added")
        for i in list_videos:
            print (i, list_videos[i])

        return list_videos

#        page = requests.get(query)
#        soup = BeautifulSoup(page.content)
#        links = soup.findAll("a")
#        i = 0
#        for link in  soup.find_all("a",href=re.compile("(?<=/url\?q=)(htt.*://.*)")):
#            list_videos[str(i)] =str(  re.split(":(?=http)",link["href"].replace("/url?q=","")))
#            i = i + 1
#        for i in list_videos:
#            print (i, list_videos[i])
#        return list_videos
    
    def get_questions(self, input):
        print ("get_questions entered")
        list_of_questions = {}
    
        wb2 = load_workbook(filename ='QuestionBank.xlsm', data_only=True)
        # get questions sheet
        sheet = wb2.get_sheet_by_name('Questions')
        # 5 for primary categories, 16 for specific categories
        category = 16
        max_rows = sheet.max_row
        for i in range(1, max_rows, 1):
            content = sheet.cell(row=i, column=category).value
            # print content
            if content == input:
                # grab the filename
                filename = sheet.cell(row=i, column=8).value
                list_of_questions[str(i)] = filename

        for i in list_of_questions:
            print (i, list_of_questions[i])
        return list_of_questions

# example of how to use the above class 
# import getContent from generate

# myContent = getContent()
# myContent.get_smry("test.txt")
# myContent.get_videos("Work and Kinetic Energy")
# myContent.get_questions("Work, Power, and Energy")


